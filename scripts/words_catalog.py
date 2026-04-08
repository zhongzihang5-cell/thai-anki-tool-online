#!/usr/bin/env python3
"""
Read Excel word list + Anki collection (SQLite), merge status.
Columns 0–7: 泰文…已熟记可删除; column 9: 备注-是否已判断过 (=1 则不算待录入).
Anki: 每词通常 1 条 note；flds 为 泰文\\x1f中文\\x1f音标\\x1f词条泰音\\x1f词条中音\\x1f
之后每 5 字段一组例句：泰例句\\x1fIPA\\x1f[sound:]\\x1f中译\\x1f[sound:]。
Stdin JSON: { "excelPath", "ankiDbPath", "sheetName"? }
sheetName defaults to 150篇集合; paths default from argv[1] JSON file if keys missing.
"""

from __future__ import annotations

import json
import os
import re
import shutil
import sqlite3
import sys
import tempfile
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print(
        json.dumps({"error": "需要安装 openpyxl: pip install -r scripts/requirements-words.txt"}),
        file=sys.stderr,
    )
    sys.exit(2)


def load_defaults(path: str | None) -> dict[str, str]:
    if not path:
        return {}
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except OSError:
        return {}


def parse_bool(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    return s in ("true", "1", "yes", "y", "是")


def remark_judged_done(v: Any) -> bool:
    """列9 备注-是否已判断过：为 1 视为已判断过。"""
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return int(v) == 1
    s = str(v).strip()
    if s == "":
        return False
    try:
        return int(float(s)) == 1
    except (TypeError, ValueError):
        return s == "1"


def strip_html(s: str) -> str:
    return re.sub(r"<[^>]+>", "\n", s)


def thai_key(s: str) -> str:
    """Match key: strip + collapse internal whitespace."""
    t = strip_html(str(s or ""))
    return "".join(t.split())


def thai_key_from_first_field(parts: list[str]) -> str | None:
    """flds 分割后第 0 字段为泰文；HTML 仅去标签后无文本则 None。"""
    raw = (parts[0] if parts else "").strip()
    if not raw:
        return None
    if raw.startswith("<"):
        key = thai_key(raw)
        return key if key else None
    return thai_key(raw) or None


def _is_thai_example_field(text: str) -> bool:
    """泰文例句槽：有实质文本，且非单独的 Anki [sound:…] 占位。"""
    t = strip_html(str(text)).strip()
    if not t:
        return False
    if re.match(r"^\[sound:", t, re.IGNORECASE):
        return False
    return True


def count_example_slots(parts: list[str]) -> int:
    """
    词条区：0 泰文、1 中文、2 音标、3–4 词条泰/中音频。

    例句区：泰文例句始终在 index 5, 10, 15, …（即 5+5k）。
    - 第 1 组 (5–9)：泰例句、IPA、[sound:]、中译、[sound:]
    - 第 2 组起 (10–14, 15–19, …)：泰例句、[sound:]、IPA、中译、[sound:]
    （组内顺序不同，但「泰文例句」永远是每组的第一个字段，故步长 5 即可。）
    """
    slot_size = 5
    start = 5
    n = 0
    k = 0
    while start + k * slot_size < len(parts):
        idx = start + k * slot_size
        raw = parts[idx] if idx < len(parts) else ""
        if _is_thai_example_field(raw):
            n += 1
        k += 1
    return n


def parse_note_flds(flds: str) -> tuple[str | None, int]:
    """返回 (泰文归一化键, 例句数)。无法识别泰文时 (None, 0)。"""
    if not flds:
        return None, 0
    parts = flds.split("\x1f")
    wkey = thai_key_from_first_field(parts)
    if not wkey:
        return None, 0
    return wkey, count_example_slots(parts)


def load_anki_map(db_path: str) -> dict[str, int]:
    """
    仅读取一个 collection.anki2：SELECT flds FROM notes。
    返回 { 泰文键: 例句数量 }；同一键多条 note 时取例句数的较大值。
    Copy DB 到临时文件再读，避免锁库。
    """
    p = Path(db_path).expanduser().resolve()
    if not p.is_file():
        print(f"ANKI_DEBUG 文件不存在: {p}", file=sys.stderr)
        return {}
    if p.suffix.lower() != ".anki2":
        print(f"ANKI_DEBUG 警告: 期望 collection.anki2，当前扩展名: {p.suffix!r} ({p})", file=sys.stderr)
    if p.name.lower() != "collection.anki2":
        print(f"ANKI_DEBUG 警告: 文件名应为 collection.anki2，当前: {p.name!r}", file=sys.stderr)

    tmp = tempfile.NamedTemporaryFile(suffix=".anki2", delete=False)
    tmp_path = tmp.name
    tmp.close()
    try:
        shutil.copy2(p, tmp_path)
        conn = sqlite3.connect(tmp_path)
    except (OSError, sqlite3.Error) as e:
        print(f"ANKI_DEBUG 打开失败: {e}", file=sys.stderr)
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        return {}
    try:
        (n_rows,) = conn.execute("SELECT COUNT(*) FROM notes").fetchone()
        cur = conn.execute("SELECT flds FROM notes")
        out: dict[str, int] = {}
        skipped = 0
        for (flds,) in cur:
            wkey, n_ex = parse_note_flds(flds)
            if wkey is None:
                skipped += 1
                continue
            prev = out.get(wkey, 0)
            out[wkey] = max(prev, n_ex)

        total_examples = sum(out.values())
        print(
            f"ANKI_DEBUG 仅连接 1 个库: {p}\n"
            f"ANKI_DEBUG notes 行数={n_rows} 有效词数={len(out)} 跳过={skipped} 例句字段合计={total_examples}",
            file=sys.stderr,
        )
        top = sorted(out.items(), key=lambda x: (-x[1], x[0]))[:10]
        print("ANKI_DEBUG 例句数 Top10（按例句数降序）:", file=sys.stderr)
        for w, c in top:
            print(f"ANKI_DEBUG   {w!r} -> {c}", file=sys.stderr)

        return out
    finally:
        conn.close()
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


def required_examples(need2: bool, need3: bool) -> int:
    """列6=True→3；列5=True→2；列5/6皆否→默认 1。"""
    if need3:
        return 3
    if need2:
        return 2
    return 1


def compute_status(
    *,
    need_entry: bool,
    need2: bool,
    need3: bool,
    deletable: bool,
    remark_judged: bool,
    in_anki: bool,
    anki_count: int,
) -> tuple[str, str]:
    """Returns (statusId, statusLabel)."""
    req = required_examples(need2, need3)

    if deletable:
        return "deletable", "可删除"

    if not in_anki:
        if need_entry:
            if remark_judged:
                return "judged", "已判断"
            return "pending", "待录入"
        return "idle", "无需录入"

    if anki_count < req:
        return "supplement", "需补充例句"

    return "done", "已完成"


def main() -> None:
    cfg_file = sys.argv[1] if len(sys.argv) > 1 else None
    defaults = load_defaults(cfg_file)

    try:
        payload = json.load(sys.stdin)
    except json.JSONDecodeError as e:
        print(json.dumps({"error": f"stdin json: {e}"}, ensure_ascii=False), file=sys.stderr)
        sys.exit(2)

    excel_path = payload.get("excelPath") or defaults.get("excelPath")
    anki_path = payload.get("ankiDbPath") or defaults.get("ankiDbPath")
    sheet_name = payload.get("sheetName") or defaults.get("sheetName") or "150篇集合"

    if not excel_path:
        print(json.dumps({"error": "缺少 excelPath"}, ensure_ascii=False), file=sys.stderr)
        sys.exit(2)

    xp = Path(excel_path)
    if not xp.is_file():
        print(
            json.dumps({"error": f"Excel 文件不存在: {excel_path}"}, ensure_ascii=False),
            file=sys.stderr,
        )
        sys.exit(2)

    anki_map = load_anki_map(anki_path) if anki_path else {}

    wb = load_workbook(xp, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(
            json.dumps(
                {"error": f"找不到工作表「{sheet_name}」，现有: {wb.sheetnames}"},
                ensure_ascii=False,
            ),
            file=sys.stderr,
        )
        sys.exit(2)
    ws = wb[sheet_name]

    words: list[dict[str, Any]] = []
    stats = {"pending": 0, "supplement": 0, "deletable": 0, "done": 0, "judged": 0}

    for row in ws.iter_rows(min_row=2, max_col=10, values_only=True):
        if not row or row[0] is None:
            continue
        thai = str(row[0]).strip()
        if not thai or thai.lower() == "泰文":
            continue

        ipa = "" if row[1] is None else str(row[1]).strip()
        zh = "" if row[2] is None else str(row[2]).strip()
        freq_raw = row[3]
        try:
            freq = int(float(freq_raw)) if freq_raw is not None and str(freq_raw).strip() != "" else 0
        except (TypeError, ValueError):
            freq = 0

        need_entry = parse_bool(row[4]) if len(row) > 4 else False
        need2 = parse_bool(row[5]) if len(row) > 5 else False
        need3 = parse_bool(row[6]) if len(row) > 6 else False
        deletable = parse_bool(row[7]) if len(row) > 7 else False
        remark_judged = remark_judged_done(row[9]) if len(row) > 9 else False

        key = thai_key(thai)
        in_anki = key in anki_map
        anki_count = anki_map.get(key, 0)
        req = required_examples(need2, need3)

        sid, label = compute_status(
            need_entry=need_entry,
            need2=need2,
            need3=need3,
            deletable=deletable,
            remark_judged=remark_judged,
            in_anki=in_anki,
            anki_count=anki_count,
        )

        if sid == "pending":
            stats["pending"] += 1
        elif sid == "supplement":
            stats["supplement"] += 1
        elif sid == "deletable":
            stats["deletable"] += 1
        elif sid == "done":
            stats["done"] += 1
        elif sid == "judged":
            stats["judged"] += 1

        need_label = str(req)

        words.append(
            {
                "thai": thai,
                "ipa": ipa,
                "chinese": zh,
                "frequency": freq,
                "needEntry": need_entry,
                "need2": need2,
                "need3": need3,
                "deletable": deletable,
                "remarkJudged": remark_judged,
                "requiredExamples": req,
                "requiredLabel": need_label,
                "inAnki": in_anki,
                "ankiNoteCount": anki_count,
                "ankiExampleCount": anki_count,
                "status": sid,
                "statusLabel": label,
            }
        )

    wb.close()

    print(json.dumps({"stats": stats, "words": words}, ensure_ascii=False))


if __name__ == "__main__":
    main()
